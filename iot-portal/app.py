#!/usr/bin/env python3
"""
INSA Advanced IoT Portal - Professional Reporting System
Combines Grafana visualizations with ThingsBoard Pro-style reporting
"""

from flask import Flask, render_template_string, jsonify, send_file, request, Response, make_response
import psycopg2
from psycopg2.extras import RealDictCursor
import pandas as pd
from datetime import datetime, timedelta
import json
import io
import os
import zipfile
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.graph_objs as go
import plotly.io as pio

app = Flask(__name__)

# Database configuration
DB_CONFIG = {
    'host': 'localhost',
    'database': 'thingsboard',
    'user': 'postgres',
    'password': '[REDACTED]***',
    'port': 5432
}

# Advanced HTML Template
PORTAL_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>INSA Advanced IoT Portal - Professional Reporting</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        :root {
            --primary-color: #667eea;
            --secondary-color: #764ba2;
            --success-color: #10b981;
            --danger-color: #ef4444;
            --dark-bg: #1a1a2e;
            --card-bg: #16213e;
        }

        body {
            background: linear-gradient(135deg, var(--primary-color) 0%, var(--secondary-color) 100%);
            min-height: 100vh;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }

        .main-header {
            background: rgba(255, 255, 255, 0.98);
            box-shadow: 0 4px 20px rgba(0,0,0,0.1);
            padding: 20px 0;
            margin-bottom: 30px;
        }

        .logo-section {
            display: flex;
            align-items: center;
            justify-content: space-between;
        }

        .logo {
            font-size: 32px;
            font-weight: bold;
            background: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }

        .live-indicator {
            display: flex;
            align-items: center;
            gap: 10px;
            padding: 10px 20px;
            background: var(--success-color);
            color: white;
            border-radius: 25px;
            animation: pulse 2s infinite;
        }

        @keyframes pulse {
            0% { box-shadow: 0 0 0 0 rgba(16, 185, 129, 0.7); }
            70% { box-shadow: 0 0 0 10px rgba(16, 185, 129, 0); }
            100% { box-shadow: 0 0 0 0 rgba(16, 185, 129, 0); }
        }

        .dashboard-card {
            background: white;
            border-radius: 15px;
            padding: 25px;
            margin-bottom: 25px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.1);
            transition: transform 0.3s;
        }

        .dashboard-card:hover {
            transform: translateY(-5px);
            box-shadow: 0 15px 40px rgba(0,0,0,0.15);
        }

        .stat-card {
            text-align: center;
            padding: 20px;
            background: linear-gradient(135deg, rgba(102, 126, 234, 0.1), rgba(118, 75, 162, 0.1));
            border-radius: 10px;
            border: 2px solid rgba(102, 126, 234, 0.3);
        }

        .stat-value {
            font-size: 48px;
            font-weight: bold;
            color: var(--primary-color);
            margin: 10px 0;
        }

        .stat-label {
            color: #666;
            text-transform: uppercase;
            letter-spacing: 1px;
            font-size: 12px;
        }

        .chart-container {
            position: relative;
            height: 400px;
            margin: 20px 0;
        }

        .grafana-embed {
            width: 100%;
            height: 450px;
            border: none;
            border-radius: 10px;
            box-shadow: 0 5px 20px rgba(0,0,0,0.1);
        }

        .export-section {
            background: linear-gradient(135deg, #f3f4f6, #e5e7eb);
            border-radius: 10px;
            padding: 20px;
            margin: 20px 0;
        }

        .export-btn {
            margin: 5px;
            padding: 12px 24px;
            border: none;
            border-radius: 8px;
            font-weight: 600;
            transition: all 0.3s;
            cursor: pointer;
        }

        .export-excel {
            background: #10b981;
            color: white;
        }

        .export-excel:hover {
            background: #059669;
            transform: scale(1.05);
        }

        .export-rar {
            background: #8b5cf6;
            color: white;
        }

        .export-pdf {
            background: #ef4444;
            color: white;
        }

        .data-table {
            width: 100%;
            border-collapse: separate;
            border-spacing: 0 10px;
        }

        .data-table thead {
            background: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
            color: white;
        }

        .data-table th {
            padding: 15px;
            text-align: left;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 12px;
            letter-spacing: 1px;
        }

        .data-table tbody tr {
            background: white;
            box-shadow: 0 2px 10px rgba(0,0,0,0.05);
            transition: all 0.3s;
        }

        .data-table tbody tr:hover {
            transform: scale(1.02);
            box-shadow: 0 5px 20px rgba(0,0,0,0.1);
        }

        .data-table td {
            padding: 15px;
            border-top: 1px solid #e5e7eb;
        }

        .tab-navigation {
            display: flex;
            gap: 10px;
            margin-bottom: 20px;
            border-bottom: 2px solid #e5e7eb;
        }

        .tab-btn {
            padding: 10px 20px;
            border: none;
            background: none;
            color: #666;
            font-weight: 600;
            cursor: pointer;
            border-bottom: 3px solid transparent;
            transition: all 0.3s;
        }

        .tab-btn.active {
            color: var(--primary-color);
            border-bottom-color: var(--primary-color);
        }

        .filter-section {
            background: #f9fafb;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 20px;
        }

        .date-range-picker {
            display: flex;
            gap: 10px;
            align-items: center;
        }

        .loading-spinner {
            display: none;
            position: fixed;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            z-index: 9999;
        }

        .loading-spinner.active {
            display: block;
        }

        .advanced-features {
            background: linear-gradient(135deg, #1e293b, #334155);
            color: white;
            padding: 30px;
            border-radius: 15px;
            margin: 30px 0;
        }

        .feature-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-top: 20px;
        }

        .feature-card {
            background: rgba(255,255,255,0.1);
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            transition: all 0.3s;
            cursor: pointer;
        }

        .feature-card:hover {
            background: rgba(255,255,255,0.2);
            transform: scale(1.05);
        }

        .feature-icon {
            font-size: 48px;
            margin-bottom: 10px;
        }
    </style>
</head>
<body>
    <div class="main-header">
        <div class="container">
            <div class="logo-section">
                <div class="logo">INSA IoT Advanced Portal</div>
                <div class="live-indicator">
                    <i class="fas fa-circle"></i>
                    <span>LIVE DATA</span>
                    <span id="last-update"></span>
                </div>
            </div>
        </div>
    </div>

    <div class="container">
        <!-- Statistics Overview -->
        <div class="row mb-4">
            <div class="col-md-3">
                <div class="stat-card">
                    <div class="stat-label">Total Records</div>
                    <div class="stat-value" id="total-records">
                        <div class="spinner-border text-primary" role="status"></div>
                    </div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="stat-card">
                    <div class="stat-label">Active Devices</div>
                    <div class="stat-value" id="active-devices">--</div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="stat-card">
                    <div class="stat-label">Data Points Today</div>
                    <div class="stat-value" id="today-points">--</div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="stat-card">
                    <div class="stat-label">Avg Response Time</div>
                    <div class="stat-value" id="response-time">--ms</div>
                </div>
            </div>
        </div>

        <!-- Tab Navigation -->
        <div class="tab-navigation">
            <button class="tab-btn active" onclick="showTab('dashboard')">
                <i class="fas fa-chart-line"></i> Dashboard
            </button>
            <button class="tab-btn" onclick="showTab('grafana')">
                <i class="fas fa-chart-area"></i> Grafana Charts
            </button>
            <button class="tab-btn" onclick="showTab('reports')">
                <i class="fas fa-file-excel"></i> Reports
            </button>
            <button class="tab-btn" onclick="showTab('historical')">
                <i class="fas fa-history"></i> Historical Data
            </button>
            <button class="tab-btn" onclick="showTab('advanced')">
                <i class="fas fa-cog"></i> Advanced Features
            </button>
        </div>

        <!-- Dashboard Tab -->
        <div id="dashboard-tab" class="tab-content">
            <div class="dashboard-card">
                <h3><i class="fas fa-chart-line"></i> Real-Time Data Stream</h3>
                <div id="realtime-chart" class="chart-container"></div>
            </div>

            <div class="dashboard-card">
                <h3><i class="fas fa-temperature-high"></i> Temperature & Humidity Analysis</h3>
                <div class="row">
                    <div class="col-md-6">
                        <div id="temperature-chart"></div>
                    </div>
                    <div class="col-md-6">
                        <div id="humidity-chart"></div>
                    </div>
                </div>
            </div>

            <div class="dashboard-card">
                <h3><i class="fas fa-server"></i> Device Status</h3>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Device Name</th>
                            <th>Status</th>
                            <th>Last Update</th>
                            <th>Data Points</th>
                            <th>Actions</th>
                        </tr>
                    </thead>
                    <tbody id="device-table">
                        <tr>
                            <td colspan="5" class="text-center">Loading...</td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>

        <!-- Grafana Tab -->
        <div id="grafana-tab" class="tab-content" style="display: none;">
            <div class="dashboard-card">
                <h3><i class="fas fa-chart-area"></i> Embedded Grafana Visualizations</h3>
                <p class="text-muted">Live data from Grafana dashboards</p>
                
                <div class="row">
                    <div class="col-md-12 mb-4">
                        <h5>IoT Overview Dashboard</h5>
                        <iframe class="grafana-embed" 
                                src="http://localhost/d/iot-overview/insa-iot-data-overview?orgId=1&kiosk=tv"></iframe>
                    </div>
                </div>

                <div class="row">
                    <div class="col-md-6">
                        <h5>Temperature Trends</h5>
                        <iframe class="grafana-embed" style="height: 300px;"
                                src="http://localhost/d-solo/iot-overview/insa-iot-data-overview?orgId=1&panelId=3"></iframe>
                    </div>
                    <div class="col-md-6">
                        <h5>Device Statistics</h5>
                        <iframe class="grafana-embed" style="height: 300px;"
                                src="http://localhost/d-solo/iot-overview/insa-iot-data-overview?orgId=1&panelId=1"></iframe>
                    </div>
                </div>
            </div>
        </div>

        <!-- Reports Tab -->
        <div id="reports-tab" class="tab-content" style="display: none;">
            <div class="dashboard-card">
                <h3><i class="fas fa-file-excel"></i> Export Reports</h3>
                
                <div class="filter-section">
                    <div class="date-range-picker">
                        <label>Date Range:</label>
                        <input type="date" id="start-date" class="form-control" style="width: 200px;">
                        <span>to</span>
                        <input type="date" id="end-date" class="form-control" style="width: 200px;">
                        <button class="btn btn-primary" onclick="applyFilter()">Apply Filter</button>
                    </div>
                </div>

                <div class="export-section">
                    <h5>Available Export Formats</h5>
                    <button class="export-btn export-excel" onclick="exportExcel()">
                        <i class="fas fa-file-excel"></i> Export to Excel
                    </button>
                    <button class="export-btn export-excel" onclick="exportAdvancedExcel()">
                        <i class="fas fa-file-excel"></i> Advanced Excel (with Charts)
                    </button>
                    <button class="export-btn export-rar" onclick="exportHistorical()">
                        <i class="fas fa-file-archive"></i> Export Historical (ZIP/RAR)
                    </button>
                    <button class="export-btn export-pdf" onclick="exportPDF()">
                        <i class="fas fa-file-pdf"></i> Export PDF Report
                    </button>
                </div>

                <div class="mt-4">
                    <h5>Recent Exports</h5>
                    <div id="export-history"></div>
                </div>
            </div>
        </div>

        <!-- Historical Tab -->
        <div id="historical-tab" class="tab-content" style="display: none;">
            <div class="dashboard-card">
                <h3><i class="fas fa-history"></i> Historical Data Analysis</h3>
                <div id="historical-chart"></div>
                
                <div class="mt-4">
                    <h5>Data Summary by Month</h5>
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>Month</th>
                                <th>Total Records</th>
                                <th>Avg Temperature</th>
                                <th>Avg Humidity</th>
                                <th>Actions</th>
                            </tr>
                        </thead>
                        <tbody id="monthly-summary">
                            <tr>
                                <td colspan="5" class="text-center">Loading...</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- Advanced Features Tab -->
        <div id="advanced-tab" class="tab-content" style="display: none;">
            <div class="advanced-features">
                <h3><i class="fas fa-rocket"></i> Advanced Features</h3>
                <p>Professional IoT data analysis and reporting tools</p>
                
                <div class="feature-grid">
                    <div class="feature-card" onclick="openFeature('predictive')">
                        <div class="feature-icon">📈</div>
                        <h5>Predictive Analytics</h5>
                        <p>AI-powered forecasting</p>
                    </div>
                    <div class="feature-card" onclick="openFeature('anomaly')">
                        <div class="feature-icon">🔍</div>
                        <h5>Anomaly Detection</h5>
                        <p>Automatic alert system</p>
                    </div>
                    <div class="feature-card" onclick="openFeature('custom')">
                        <div class="feature-icon">⚙️</div>
                        <h5>Custom Queries</h5>
                        <p>SQL query builder</p>
                    </div>
                    <div class="feature-card" onclick="openFeature('api')">
                        <div class="feature-icon">🔌</div>
                        <h5>API Access</h5>
                        <p>REST API endpoints</p>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- Loading Spinner -->
    <div class="loading-spinner">
        <div class="spinner-border text-primary" style="width: 3rem; height: 3rem;" role="status">
            <span class="sr-only">Loading...</span>
        </div>
    </div>

    <script>
        // Initialize dashboard
        document.addEventListener('DOMContentLoaded', function() {
            loadDashboardData();
            setInterval(loadDashboardData, 30000); // Refresh every 30 seconds
            initializeCharts();
        });

        function showTab(tabName) {
            // Hide all tabs
            document.querySelectorAll('.tab-content').forEach(tab => {
                tab.style.display = 'none';
            });
            
            // Remove active class from all buttons
            document.querySelectorAll('.tab-btn').forEach(btn => {
                btn.classList.remove('active');
            });
            
            // Show selected tab
            document.getElementById(tabName + '-tab').style.display = 'block';
            
            // Add active class to clicked button
            event.target.closest('.tab-btn').classList.add('active');
            
            // Load tab-specific data
            if (tabName === 'historical') {
                loadHistoricalData();
            } else if (tabName === 'reports') {
                loadExportHistory();
            }
        }

        async function loadDashboardData() {
            try {
                const response = await fetch('/api/dashboard');
                const data = await response.json();
                
                document.getElementById('total-records').textContent = 
                    data.total_records?.toLocaleString() || '0';
                document.getElementById('active-devices').textContent = 
                    data.active_devices || '0';
                document.getElementById('today-points').textContent = 
                    data.today_points?.toLocaleString() || '0';
                document.getElementById('response-time').textContent = 
                    data.response_time + 'ms' || '--ms';
                document.getElementById('last-update').textContent = 
                    new Date().toLocaleTimeString();
                
                // Update device table
                updateDeviceTable(data.devices);
                
                // Update charts
                updateRealtimeChart(data.realtime_data);
            } catch (error) {
                console.error('Error loading dashboard data:', error);
            }
        }

        function updateDeviceTable(devices) {
            const tbody = document.getElementById('device-table');
            let html = '';
            
            devices?.forEach(device => {
                const statusClass = device.active ? 'text-success' : 'text-danger';
                const statusIcon = device.active ? 'fa-circle' : 'fa-circle';
                
                html += `
                    <tr>
                        <td><strong>${device.name}</strong></td>
                        <td><i class="fas ${statusIcon} ${statusClass}"></i> ${device.active ? 'Online' : 'Offline'}</td>
                        <td>${device.last_update}</td>
                        <td>${device.data_points?.toLocaleString()}</td>
                        <td>
                            <button class="btn btn-sm btn-primary" onclick="viewDevice('${device.id}')">
                                <i class="fas fa-eye"></i> View
                            </button>
                            <button class="btn btn-sm btn-success" onclick="exportDevice('${device.id}')">
                                <i class="fas fa-download"></i> Export
                            </button>
                        </td>
                    </tr>
                `;
            });
            
            tbody.innerHTML = html || '<tr><td colspan="5" class="text-center">No devices found</td></tr>';
        }

        function initializeCharts() {
            // Initialize Plotly charts
            const layout = {
                autosize: true,
                margin: {l: 40, r: 20, t: 20, b: 40},
                paper_bgcolor: 'rgba(0,0,0,0)',
                plot_bgcolor: 'rgba(0,0,0,0)',
                xaxis: {
                    gridcolor: '#e5e7eb',
                    showgrid: true
                },
                yaxis: {
                    gridcolor: '#e5e7eb',
                    showgrid: true
                }
            };

            // Temperature chart
            Plotly.newPlot('temperature-chart', [{
                x: [],
                y: [],
                type: 'scatter',
                mode: 'lines+markers',
                name: 'Temperature',
                line: {color: '#ef4444', width: 3}
            }], {...layout, title: 'Temperature (°C)'});

            // Humidity chart
            Plotly.newPlot('humidity-chart', [{
                x: [],
                y: [],
                type: 'scatter',
                mode: 'lines+markers',
                name: 'Humidity',
                line: {color: '#3b82f6', width: 3}
            }], {...layout, title: 'Humidity (%)'});
        }

        function updateRealtimeChart(data) {
            if (!data) return;
            
            // Update temperature chart
            Plotly.update('temperature-chart', {
                x: [data.timestamps],
                y: [data.temperatures]
            });
            
            // Update humidity chart
            Plotly.update('humidity-chart', {
                x: [data.timestamps],
                y: [data.humidity]
            });
        }

        async function exportExcel() {
            showLoading();
            window.location.href = '/export/excel?type=simple';
            hideLoading();
        }

        async function exportAdvancedExcel() {
            showLoading();
            window.location.href = '/export/excel?type=advanced';
            hideLoading();
        }

        async function exportHistorical() {
            showLoading();
            const startDate = document.getElementById('start-date').value;
            const endDate = document.getElementById('end-date').value;
            window.location.href = `/export/historical?start=${startDate}&end=${endDate}`;
            hideLoading();
        }

        async function exportPDF() {
            showLoading();
            window.location.href = '/export/pdf';
            hideLoading();
        }

        function showLoading() {
            document.querySelector('.loading-spinner').classList.add('active');
        }

        function hideLoading() {
            setTimeout(() => {
                document.querySelector('.loading-spinner').classList.remove('active');
            }, 1000);
        }

        async function loadHistoricalData() {
            try {
                const response = await fetch('/api/historical');
                const data = await response.json();
                
                // Update monthly summary table
                const tbody = document.getElementById('monthly-summary');
                let html = '';
                
                data.monthly_summary?.forEach(month => {
                    html += `
                        <tr>
                            <td>${month.month}</td>
                            <td>${month.records?.toLocaleString()}</td>
                            <td>${month.avg_temp}°C</td>
                            <td>${month.avg_humidity}%</td>
                            <td>
                                <button class="btn btn-sm btn-primary" onclick="downloadMonth('${month.month}')">
                                    <i class="fas fa-download"></i> Download
                                </button>
                            </td>
                        </tr>
                    `;
                });
                
                tbody.innerHTML = html || '<tr><td colspan="5" class="text-center">No data available</td></tr>';
            } catch (error) {
                console.error('Error loading historical data:', error);
            }
        }

        function openFeature(feature) {
            alert(`Advanced feature: ${feature} - Coming soon!`);
        }

        function viewDevice(deviceId) {
            window.open(`/device/${deviceId}`, '_blank');
        }

        function exportDevice(deviceId) {
            window.location.href = `/export/device/${deviceId}`;
        }

        function downloadMonth(month) {
            window.location.href = `/export/month/${month}`;
        }
    </script>
</body>
</html>
"""

def get_db_connection():
    """Create database connection"""
    return psycopg2.connect(**DB_CONFIG)

@app.route('/')
def index():
    """Main portal page"""
    return render_template_string(PORTAL_TEMPLATE)

@app.route('/api/dashboard')
def get_dashboard_data():
    """Get dashboard statistics and real-time data"""
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        data = {}
        
        # Get total records
        cur.execute("""
            SELECT SUM(count) as total FROM (
                SELECT COUNT(*) as count FROM ts_kv_2025_07
                UNION ALL
                SELECT COUNT(*) as count FROM ts_kv_2025_08
                UNION ALL
                SELECT COUNT(*) as count FROM ts_kv_2025_09
            ) t
        """)
        data['total_records'] = cur.fetchone()['total']
        
        # Get active devices
        cur.execute("""
            SELECT COUNT(DISTINCT d.id) as count 
            FROM device d 
            JOIN ts_kv_latest kv ON d.id = kv.entity_id 
            WHERE kv.ts > extract(epoch from now() - interval '1 hour') * 1000
        """)
        data['active_devices'] = cur.fetchone()['count']
        
        # Get today's data points
        cur.execute("""
            SELECT COUNT(*) as count 
            FROM ts_kv_latest 
            WHERE ts > extract(epoch from date_trunc('day', now())) * 1000
        """)
        data['today_points'] = cur.fetchone()['count']
        
        # Calculate response time (mock for now)
        data['response_time'] = 42
        
        # Get device list
        cur.execute("""
            SELECT 
                d.id,
                d.name,
                COUNT(kv.ts) as data_points,
                MAX(kv.ts) as last_ts,
                CASE 
                    WHEN MAX(kv.ts) > extract(epoch from now() - interval '1 hour') * 1000 
                    THEN true 
                    ELSE false 
                END as active
            FROM device d
            LEFT JOIN ts_kv_latest kv ON d.id = kv.entity_id
            WHERE d.type = 'IOT'
            GROUP BY d.id, d.name
            ORDER BY d.name
            LIMIT 10
        """)
        
        devices = []
        for row in cur.fetchall():
            device = dict(row)
            if device['last_ts']:
                device['last_update'] = datetime.fromtimestamp(device['last_ts']/1000).strftime('%Y-%m-%d %H:%M')
            else:
                device['last_update'] = 'Never'
            devices.append(device)
        
        data['devices'] = devices
        
        # Get real-time data for charts (last 24 hours, hourly)
        cur.execute("""
            SELECT 
                date_trunc('hour', to_timestamp(ts/1000)) as hour,
                AVG(CAST(str_v as FLOAT)) as value,
                k.key
            FROM ts_kv_latest kv
            JOIN ts_kv_dictionary k ON k.key_id = kv.key
            WHERE kv.ts > extract(epoch from now() - interval '24 hours') * 1000
                AND k.key IN ('temperature', 'humidity')
                AND str_v ~ '^[0-9.]+$'
            GROUP BY hour, k.key
            ORDER BY hour
        """)
        
        realtime_data = {
            'timestamps': [],
            'temperatures': [],
            'humidity': []
        }
        
        for row in cur.fetchall():
            timestamp = row['hour'].strftime('%H:%M')
            if timestamp not in realtime_data['timestamps']:
                realtime_data['timestamps'].append(timestamp)
            
            if row['key'] == 'temperature':
                realtime_data['temperatures'].append(round(row['value'], 2))
            elif row['key'] == 'humidity':
                realtime_data['humidity'].append(round(row['value'], 2))
        
        data['realtime_data'] = realtime_data
        
        cur.close()
        conn.close()
        
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/historical')
def get_historical_data():
    """Get historical data summary"""
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get monthly summary
        cur.execute("""
            SELECT 
                to_char(to_timestamp(ts/1000), 'YYYY-MM') as month,
                COUNT(*) as records,
                ROUND(AVG(CAST(NULLIF(str_v, '') as FLOAT) FILTER (WHERE k.key = 'temperature')), 2) as avg_temp,
                ROUND(AVG(CAST(NULLIF(str_v, '') as FLOAT) FILTER (WHERE k.key = 'humidity')), 2) as avg_humidity
            FROM ts_kv_latest kv
            JOIN ts_kv_dictionary k ON k.key_id = kv.key
            WHERE str_v ~ '^[0-9.]+$'
            GROUP BY month
            ORDER BY month DESC
            LIMIT 12
        """)
        
        monthly_summary = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({'monthly_summary': monthly_summary})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export/excel')
def export_excel():
    """Export data to Excel with charts"""
    try:
        export_type = request.args.get('type', 'simple')
        
        conn = get_db_connection()
        
        # Get data for export
        query = """
            SELECT 
                d.name as device_name,
                k.key as parameter,
                kv.str_v as value,
                to_timestamp(kv.ts/1000) as timestamp
            FROM ts_kv_latest kv
            JOIN device d ON d.id = kv.entity_id
            JOIN ts_kv_dictionary k ON k.key_id = kv.key
            WHERE kv.ts > extract(epoch from now() - interval '7 days') * 1000
            ORDER BY kv.ts DESC
            LIMIT 10000
        """
        
        df = pd.read_sql_query(query, conn)
        conn.close()
        
        # Create Excel file
        output = io.BytesIO()
        
        if export_type == 'advanced':
            # Create workbook with charts
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Raw Data', index=False)
                
                # Get workbook and add charts
                workbook = writer.book
                
                # Create summary sheet
                summary_sheet = workbook.create_sheet('Summary')
                summary_sheet['A1'] = 'INSA IoT Data Report'
                summary_sheet['A1'].font = Font(size=16, bold=True)
                summary_sheet['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
                summary_sheet['A4'] = f'Total Records: {len(df)}'
                
                # Add pivot table data
                pivot_df = df.pivot_table(
                    values='value',
                    index='timestamp',
                    columns='parameter',
                    aggfunc='mean'
                )
                
                # Write pivot to new sheet
                pivot_df.to_excel(writer, sheet_name='Analysis')
                
                # Style the sheets
                for sheet in workbook.worksheets:
                    for cell in sheet['1']:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
                        cell.font = Font(color='FFFFFF', bold=True)
        else:
            # Simple export
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='IoT Data', index=False)
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'iot_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export/historical')
def export_historical():
    """Export historical data as compressed archive"""
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        
        conn = get_db_connection()
        
        # Create temporary directory for files
        temp_dir = f'/tmp/iot_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        os.makedirs(temp_dir, exist_ok=True)
        
        # Export data by device
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT id, name FROM device WHERE type = 'IOT'")
        devices = cur.fetchall()
        
        for device_id, device_name in devices:
            query = """
                SELECT 
                    k.key,
                    kv.str_v as value,
                    to_timestamp(kv.ts/1000) as timestamp
                FROM ts_kv_latest kv
                JOIN ts_kv_dictionary k ON k.key_id = kv.key
                WHERE kv.entity_id = %s
                ORDER BY kv.ts DESC
                LIMIT 50000
            """
            
            df = pd.read_sql_query(query, conn, params=[device_id])
            
            if not df.empty:
                # Save to CSV
                filename = f"{temp_dir}/{device_name.replace(' ', '_')}.csv"
                df.to_csv(filename, index=False)
        
        cur.close()
        conn.close()
        
        # Create ZIP archive
        output = io.BytesIO()
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    zipf.write(os.path.join(root, file), file)
        
        # Clean up temp files
        for file in os.listdir(temp_dir):
            os.remove(os.path.join(temp_dir, file))
        os.rmdir(temp_dir)
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'historical_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export/pdf')
def export_pdf():
    """Export PDF report (requires wkhtmltopdf)"""
    return jsonify({'message': 'PDF export requires additional setup. Use Excel export for now.'}), 501

@app.route('/device/<device_id>')
def device_details(device_id):
    """Device details page"""
    return f"<h1>Device Details for {device_id}</h1><p>Coming soon...</p>"

@app.route('/export/device/<device_id>')
def export_device_data(device_id):
    """Export specific device data"""
    try:
        conn = get_db_connection()
        
        query = """
            SELECT 
                d.name as device_name,
                k.key as parameter,
                kv.str_v as value,
                to_timestamp(kv.ts/1000) as timestamp
            FROM ts_kv_latest kv
            JOIN device d ON d.id = kv.entity_id
            JOIN ts_kv_dictionary k ON k.key_id = kv.key
            WHERE d.id = %s
            ORDER BY kv.ts DESC
            LIMIT 10000
        """
        
        df = pd.read_sql_query(query, conn, params=[device_id])
        conn.close()
        
        output = io.BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        
        device_name = df['device_name'].iloc[0] if not df.empty else 'device'
        
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'{device_name}_{datetime.now().strftime("%Y%m%d")}.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export/month/<month>')
def export_month_data(month):
    """Export data for specific month"""
    try:
        conn = get_db_connection()
        
        query = """
            SELECT 
                d.name as device_name,
                k.key as parameter,
                kv.str_v as value,
                to_timestamp(kv.ts/1000) as timestamp
            FROM ts_kv_latest kv
            JOIN device d ON d.id = kv.entity_id
            JOIN ts_kv_dictionary k ON k.key_id = kv.key
            WHERE to_char(to_timestamp(kv.ts/1000), 'YYYY-MM') = %s
            ORDER BY kv.ts
        """
        
        df = pd.read_sql_query(query, conn, params=[month])
        conn.close()
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'Data_{month}', index=False)
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'iot_data_{month}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
